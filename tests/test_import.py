"""Excel import: template, partial success, slug unikallik va multi-tenant izolyatsiya."""

import io

from openpyxl import Workbook, load_workbook
from sqlalchemy import select

from app.models import Category, Product
from app.services.product_import import HEADERS

ALFA = {"X-Admin-Token": "alfa-token"}
BETA = {"X-Admin-Token": "beta-token"}

_XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def make_xlsx(rows: list[list], *, sheet_name: str = "Mahsulotlar", headers: list = HEADERS) -> bytes:
    """Sarlavha qatori + berilgan data qatorlaridan xlsx baytlarini quradi."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _upload(content: bytes, headers=ALFA, filename: str = "import.xlsx"):
    return {"file": (filename, content, _XLSX_MEDIA_TYPE)}


async def _products(db, business_id: int) -> list[Product]:
    return list(
        (
            await db.scalars(
                select(Product).where(Product.business_id == business_id).order_by(Product.id)
            )
        ).all()
    )


# --- Template ---------------------------------------------------------------


async def test_template_yuklab_olinadi(client, shops):
    response = await client.get("/api/admin/products/import/template", headers=ALFA)
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(_XLSX_MEDIA_TYPE)
    workbook = load_workbook(io.BytesIO(response.content))
    assert "Mahsulotlar" in workbook.sheetnames
    assert "Yo‘riqnoma" in workbook.sheetnames
    # Sarlavha qatori shablon ustunlariga mos.
    header_row = [cell.value for cell in workbook["Mahsulotlar"][1]]
    assert header_row == HEADERS


# --- To‘g‘ri fayl -----------------------------------------------------------


async def test_togri_fayl_hammasi_yaratiladi_va_kategoriya_ochiladi(client, shops, db):
    content = make_xlsx(
        [
            ["Yumshoq divan", "Yotoqxona", "4 500 000", "5 000 000", "izoh", "Teri", "200x90", "Oq", "SKU1", "bor"],
            ["Kreslo", "Yotoqxona", 1200000, None, None, None, None, None, None, "buyurtmaga"],
        ]
    )
    response = await client.post("/api/admin/products/import", headers=ALFA, files=_upload(content))
    assert response.status_code == 200
    body = response.json()
    assert body["created"] == 2
    assert body["skipped"] == 0
    assert body["errors"] == []

    # Yangi "Yotoqxona" kategoriyasi shu biznesda ochilgan.
    category = await db.scalar(
        select(Category).where(
            Category.business_id == shops["alfa"].id, Category.name == "Yotoqxona"
        )
    )
    assert category is not None
    assert category.is_active is True
    assert category.slug == "yotoqxona"

    # Ikkala mahsulot ham shu kategoriyaga bog‘langan.
    divan = await db.scalar(
        select(Product).where(Product.business_id == shops["alfa"].id, Product.name == "Yumshoq divan")
    )
    assert divan.category_id == category.id
    assert str(divan.price) == "4500000.00"
    assert str(divan.old_price) == "5000000.00"


# --- Aralash fayl (partial success) ----------------------------------------


async def test_aralash_fayl_valid_saqlanadi_xatolar_hisobotda(client, shops, db):
    content = make_xlsx(
        [
            ["To‘g‘ri mahsulot", "", "1000000", None, None, None, None, None, None, None],  # 2-qator OK
            ["Narxsiz", "", "", None, None, None, None, None, None, None],  # 3-qator xato
            ["Yomon narx", "", "salom", None, None, None, None, None, None, None],  # 4-qator xato
            ["", "", "500000", None, None, None, None, None, None, None],  # 5-qator: nom yo‘q
        ]
    )
    response = await client.post("/api/admin/products/import", headers=ALFA, files=_upload(content))
    assert response.status_code == 200
    body = response.json()
    assert body["created"] == 1
    assert body["skipped"] == 3

    rows_with_msgs = {err["row"]: err["message"] for err in body["errors"]}
    assert set(rows_with_msgs) == {3, 4, 5}
    assert "Narx majburiy" in rows_with_msgs[3]
    assert "noto‘g‘ri formatda" in rows_with_msgs[4]
    assert "Nomi majburiy" in rows_with_msgs[5]
    # Xabar Excel qator raqami bilan boshlanadi.
    assert rows_with_msgs[4].startswith("4-qator:")

    # Faqat valid mahsulot saqlangan.
    created = await db.scalar(
        select(Product).where(
            Product.business_id == shops["alfa"].id, Product.name == "To‘g‘ri mahsulot"
        )
    )
    assert created is not None


# --- Slug unikallik ---------------------------------------------------------


async def test_takror_nom_slug_unikallashadi(client, shops, db):
    # "Alfa divan" allaqachon bor (slug: alfa-divan). Import shu nom bilan → -2.
    content = make_xlsx(
        [
            ["Alfa divan", "", "1000000", None, None, None, None, None, None, None],
            ["Alfa divan", "", "1100000", None, None, None, None, None, None, None],
        ]
    )
    response = await client.post("/api/admin/products/import", headers=ALFA, files=_upload(content))
    assert response.status_code == 200
    assert response.json()["created"] == 2

    slugs = {product.slug for product in await _products(db, shops["alfa"].id)}
    assert {"alfa-divan", "alfa-divan-2", "alfa-divan-3"} <= slugs


# --- Narx matn → qator xatosi ----------------------------------------------


async def test_narx_matn_qator_xatosi(client, shops, db):
    content = make_xlsx(
        [["Mahsulot", "", "bepul", None, None, None, None, None, None, None]]
    )
    response = await client.post("/api/admin/products/import", headers=ALFA, files=_upload(content))
    assert response.status_code == 200
    body = response.json()
    assert body["created"] == 0
    assert body["skipped"] == 1
    assert body["errors"][0]["row"] == 2
    assert "noto‘g‘ri formatda" in body["errors"][0]["message"]


async def test_eski_narx_kichik_bolsa_xato(client, shops):
    content = make_xlsx(
        [["Mahsulot", "", "1000000", "900000", None, None, None, None, None, None]]
    )
    response = await client.post("/api/admin/products/import", headers=ALFA, files=_upload(content))
    body = response.json()
    assert body["created"] == 0
    assert "Eski narx" in body["errors"][0]["message"]


# --- 422: noto‘g‘ri fayl turi va qator limiti --------------------------------


async def test_notogri_fayl_turi_422(client, shops):
    response = await client.post(
        "/api/admin/products/import",
        headers=ALFA,
        files={"file": ("data.txt", b"bu excel emas", "text/plain")},
    )
    assert response.status_code == 422


async def test_qator_limiti_oshsa_422(client, shops):
    rows = [[f"Mahsulot {i}", "", "1000000", None, None, None, None, None, None, None] for i in range(501)]
    content = make_xlsx(rows)
    response = await client.post("/api/admin/products/import", headers=ALFA, files=_upload(content))
    assert response.status_code == 422


# --- Multi-tenant izolyatsiya -----------------------------------------------


async def test_boshqa_biznes_kategoriyasi_nomi_ozida_yangi_yaratiladi(client, shops, db):
    # "Stollar" — beta biznesining kategoriyasi. Alfa importi uni ulamaydi, o‘zida yaratadi.
    beta_category = await db.scalar(
        select(Category).where(Category.business_id == shops["beta"].id, Category.name == "Stollar")
    )
    content = make_xlsx(
        [["Alfa stol", "Stollar", "2000000", None, None, None, None, None, None, None]]
    )
    response = await client.post("/api/admin/products/import", headers=ALFA, files=_upload(content))
    assert response.status_code == 200
    assert response.json()["created"] == 1

    product = await db.scalar(
        select(Product).where(Product.business_id == shops["alfa"].id, Product.name == "Alfa stol")
    )
    alfa_category = await db.scalar(
        select(Category).where(
            Category.business_id == shops["alfa"].id, Category.name == "Stollar"
        )
    )
    assert alfa_category is not None
    assert alfa_category.id != beta_category.id  # chet biznesga ulanmagan
    assert product.category_id == alfa_category.id
