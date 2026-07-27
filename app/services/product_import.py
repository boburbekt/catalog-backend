"""Excel (.xlsx) orqali mahsulot importi: shablon generatsiyasi va fayl tahlili.

Bu modul faqat fayl bilan ishlaydi (openpyxl) — DB'ga tegmaydi. Kategoriya qidirish,
slug unikallashtirish va yozish `api/admin.py` da, tenant scoping saqlangan holda bajariladi.
"""

import io
import re
from decimal import Decimal, InvalidOperation

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.models import Availability

# xlsx aslida ZIP arxiv — magic bytes PK\x03\x04 bilan boshlanadi (images.py yondashuvi kabi).
XLSX_MAGIC = b"PK\x03\x04"
MAX_IMPORT_BYTES = 5 * 1024 * 1024
MAX_IMPORT_ROWS = 500

DATA_SHEET = "Mahsulotlar"

# Ustunlar tartibi — shablon sarlavhasi ham, tahlil ham shunga tayanadi. `*` = majburiy.
HEADERS = [
    "Nomi*",
    "Kategoriya",
    "Narx (so‘m)*",
    "Eski narx",
    "Tavsif",
    "Material",
    "O‘lchamlari",
    "Rangi",
    "SKU",
    "Mavjudlik",
]

# Mavjudlik: o‘zbekcha qiymatlar + enum qiymatlarining o‘zi ham qabul qilinadi.
AVAILABILITY_MAP = {
    "bor": Availability.IN_STOCK,
    "buyurtmaga": Availability.PREORDER,
    "tugagan": Availability.OUT_OF_STOCK,
    "in_stock": Availability.IN_STOCK,
    "preorder": Availability.PREORDER,
    "out_of_stock": Availability.OUT_OF_STOCK,
}
_AVAILABILITY_HINT = "bor / buyurtmaga / tugagan"

# Minglik ajratgichlar: oddiy bo‘shliq, no-break space va vergul ("4 500 000", "4,500,000").
_THOUSANDS = re.compile(r"[\s ,]")


class RowError(Exception):
    """Bitta qator validatsiyadan o‘tmadi — xabari o‘zbekcha, hisobotga tushadi."""


def is_blank(value: object) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def cell_str(value: object) -> str:
    """Katak qiymatini tozalangan matnga aylantiradi (bo‘sh bo‘lsa `""`)."""
    if value is None:
        return ""
    return str(value).strip()


def parse_amount(value: object) -> Decimal | None:
    """Narx katagini `Decimal`ga aylantiradi. Bo‘sh → `None`, formati buzuq → `ValueError`.

    Raqam kataklar (int/float) to‘g‘ridan-to‘g‘ri; matnda minglik ajratgichlar tozalanadi.
    """
    if is_blank(value):
        return None
    if isinstance(value, bool):  # bool int'ning kichik turi — narx sifatida rad etamiz
        raise ValueError("narx raqam emas")
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    cleaned = _THOUSANDS.sub("", str(value).strip())
    try:
        return Decimal(cleaned)
    except InvalidOperation as exc:
        raise ValueError("narx raqam emas") from exc


def parse_availability(value: object) -> Availability:
    """Mavjudlik katagini enumga map qiladi. Bo‘sh → IN_STOCK, notanish → `RowError`."""
    if is_blank(value):
        return Availability.IN_STOCK
    mapped = AVAILABILITY_MAP.get(str(value).strip().lower())
    if mapped is None:
        raise RowError(f"Mavjudlik qiymati noto‘g‘ri (ruxsat etilgan: {_AVAILABILITY_HINT})")
    return mapped


def load_rows(data: bytes) -> list[tuple[int, list]]:
    """Faylni o‘qib, `(excel_qator_raqami, 10 ta katak)` ro‘yxatini qaytaradi.

    Sarlavha qatori (1-qator) o‘tkaziladi; butunlay bo‘sh qatorlar ham tashlab yuboriladi.
    Qatorlar Excel'dagi haqiqiy raqamini saqlaydi — xato hisobotida shu ko‘rsatiladi.
    """
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        worksheet = workbook[DATA_SHEET] if DATA_SHEET in workbook.sheetnames else workbook.worksheets[0]
        rows: list[tuple[int, list]] = []
        for excel_row, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if values is None or all(is_blank(cell) for cell in values):
                continue
            cells = list(values[:10]) + [None] * (10 - len(values))
            rows.append((excel_row, cells))
        return rows
    finally:
        workbook.close()


def build_template_workbook(category_names: list[str]) -> bytes:
    """Namuna qatorli, yo‘riqnomali .xlsx shablon baytlarini qaytaradi.

    `category_names` bo‘lsa Kategoriya ustuniga dropdown (data validation) qo‘yiladi.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = DATA_SHEET
    sheet.append(HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    example_category = category_names[0] if category_names else "Divanlar"
    sheet.append(
        ["Divan Milano", example_category, 4500000, 5200000, "Yumshoq burchak divan",
         "Teri", "220x90x85 sm", "Jigarrang", "DIV-001", "bor"]
    )
    sheet.append(
        ["Yotoq Roma", example_category, 3200000, None, "Ikki kishilik karavot",
         "MDF", "160x200 sm", "Oq", "YOT-002", "buyurtmaga"]
    )

    widths = [22, 18, 15, 12, 30, 14, 16, 12, 12, 14]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    # Kategoriya dropdown: inline ro‘yxatning Excel chegarasi ~255 belgi; vergulli
    # nomlar formulani buzadi, shuning uchun ularni chetlab o‘tamiz.
    safe_names = [name for name in category_names if "," not in name]
    if safe_names:
        formula = '"' + ",".join(safe_names) + '"'
        if len(formula) <= 255:
            validation = DataValidation(type="list", formula1=formula, allow_blank=True)
            sheet.add_data_validation(validation)
            validation.add(f"B2:B{MAX_IMPORT_ROWS + 1}")

    guide = workbook.create_sheet("Yo‘riqnoma")
    guide_rows = [
        ["Ustun", "Izoh"],
        ["Nomi*", "Majburiy. Mahsulot nomi."],
        ["Kategoriya", "Ixtiyoriy. Bo‘lmasa yangi kategoriya avtomatik ochiladi."],
        ["Narx (so‘m)*", "Majburiy. Faqat raqam, 0 dan katta. Masalan: 4500000 yoki 4 500 000."],
        ["Eski narx", "Ixtiyoriy. Berilsa Narxdan katta bo‘lishi shart (chegirma ko‘rsatish uchun)."],
        ["Tavsif / Material / O‘lchamlari / Rangi / SKU", "Ixtiyoriy matn maydonlari."],
        ["Mavjudlik", f"Ruxsat etilgan qiymatlar: {_AVAILABILITY_HINT}. Bo‘sh bo‘lsa 'bor' deb olinadi."],
        ["", ""],
        ["Eslatma", "Rasm Excel orqali yuklanmaydi — importdan keyin har mahsulotga alohida qo‘shiladi."],
        ["Eslatma", "Bir faylda ko‘pi bilan 500 ta mahsulot. Namuna qatorlarni o‘chirib, o‘zingiznikini kiriting."],
    ]
    for row in guide_rows:
        guide.append(row)
    for cell in guide[1]:
        cell.font = Font(bold=True)
    guide.column_dimensions["A"].width = 42
    guide.column_dimensions["B"].width = 70

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
